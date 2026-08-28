from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import CollectionMode, ExportFormat, SalesResult, TimeSlotRecord


class BaseExcelExporter:
    title = "Sales Data Collector"

    def __init__(self, *, include_store_code: bool) -> None:
        self.include_store_code = include_store_code

    def export(self, path: str | Path, results: Sequence[SalesResult]) -> Path:
        path = Path(path)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        self._write_summary_sheet(summary_sheet, results)

        if any(result.time_slots for result in results):
            time_sheet = workbook.create_sheet("Time Sales")
            self._write_time_sheet(time_sheet, results)

        self._apply_layout(workbook)
        workbook.save(path)
        return path

    def _write_summary_sheet(self, sheet, results: Sequence[SalesResult]) -> None:
        headers = ["business_date"]
        if self.include_store_code:
            headers.append("magic_store_id")
        headers.extend(["store_name", "order_count", "sales_amount", "source_status"])
        sheet.append(headers)
        for result in results:
            row = [result.business_date.isoformat()]
            if self.include_store_code:
                row.append(result.magic_store_id)
            row.extend([result.store_name, result.order_count, result.sales_amount, result.source_status.value])
            sheet.append(row)

    def _write_time_sheet(self, sheet, results: Sequence[SalesResult]) -> None:
        headers = ["business_date"]
        if self.include_store_code:
            headers.append("magic_store_id")
        headers.extend(["store_name", "time_slot", "order_count", "sales_amount"])
        sheet.append(headers)
        for result in results:
            for slot in result.time_slots:
                row = [slot.business_date.isoformat()]
                if self.include_store_code:
                    row.append(slot.magic_store_id)
                row.extend([slot.store_name, slot.time_slot, slot.order_count, slot.sales_amount])
                sheet.append(row)

    def _apply_layout(self, workbook: Workbook) -> None:
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column_cells in sheet.columns:
                column_letter = get_column_letter(column_cells[0].column)
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


class SalesAdminExcelExporter(BaseExcelExporter):
    def __init__(self) -> None:
        super().__init__(include_store_code=False)


class DeveloperExcelExporter(BaseExcelExporter):
    def __init__(self) -> None:
        super().__init__(include_store_code=True)


def build_exporter(mode: CollectionMode) -> BaseExcelExporter:
    if mode == CollectionMode.ADMIN:
        return SalesAdminExcelExporter()
    return DeveloperExcelExporter()

