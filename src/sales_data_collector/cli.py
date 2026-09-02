from __future__ import annotations

import argparse
import calendar
import os
from datetime import date
from getpass import getpass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from .client import ServiceClient
from .collector import collect_sales, export_sales
from .excel.menu_excel_dry_run import CellPlanStatus, build_menu_excel_dry_run_plan, summarize_other_residual_by_reason
from .excel.menu_copy_writer import MenuMonthlyCopyWriter
from .excel.menu_template_profile import DAEGU_JULY_PROFILE, DAEJEON_JULY_PROFILE, MenuTemplateProfile
from .exceptions import AuthenticationError, ServiceError
from .models import CollectionMode, ExportFormat, SalesAdminDailyRecord, Store
from .production import SalesAdminDryRun, SingleDaySalesCollector, default_target_date
from .mapping.template_resolver import SalesAdminTemplateResolver
from .mapping.store_normalizer import normalize_store_name
from .mapping.menu_mapping import MenuMappingStatus, build_menu_mapping_preview

from .excel.menu_excel_dry_run import other_new_menu_inventory
from .writers.single_day_writer import SalesAdminSingleDayWriter
from .utils import date_range, parse_ymd


DEFAULT_BASE_URL = os.environ.get("COLLECTOR_BASE_URL")
DEFAULT_BRAND_IDX = os.environ.get("COLLECTOR_BRAND_IDX")
DEFAULT_BRAND_NAME = os.environ.get("COLLECTOR_BRAND_NAME")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="Sales Data Collector")
    parser.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        help="Service base URL. Set COLLECTOR_BASE_URL locally or provide this option.",
    )
    parser.add_argument("--user-id", help="Service login ID")
    parser.add_argument("--brand-idx", default=DEFAULT_BRAND_IDX)
    parser.add_argument("--brand-name", default=DEFAULT_BRAND_NAME)
    parser.add_argument("--store-idx", action="append", default=[], help="Specific store ID. Repeatable.")
    parser.add_argument("--store-name", action="append", default=[], help="Specific store name. Repeatable.")
    parser.add_argument("--all-stores", action="store_true", help="Collect all returned stores")
    parser.add_argument("--start-date", help="YYYY-MM-DD (legacy diagnostic mode)")
    parser.add_argument("--end-date", help="YYYY-MM-DD (legacy diagnostic mode)")
    parser.add_argument("--output", help="Output xlsx path (legacy diagnostic mode)")
    parser.add_argument("--production-dry-run", action="store_true", help="Preview one production workbook update without writing.")
    parser.add_argument("--production-write", action="store_true", help="Copy a workbook and write confirmed production updates.")
    parser.add_argument("--menu-monthly-preview", action="store_true", help="Preview one store's monthly menu sales without writing.")
    parser.add_argument("--menu-mapping-preview", action="store_true", help="Preview raw menu normalization and mapping without writing.")
    parser.add_argument("--menu-excel-dry-run", action="store_true", help="Plan monthly menu Excel updates without writing or saving.")
    parser.add_argument("--menu-excel-write-copy", action="store_true", help="Copy a menu workbook and safely write one monthly menu update.")
    parser.add_argument("--year", type=int, help="Collection year for monthly preview modes.")
    parser.add_argument("--month", type=int, help="Collection month for monthly preview modes.")
    parser.add_argument("--date", action="append", default=[], help="YYYY-MM-DD. Repeat for multi-date production dry-run; defaults to yesterday.")
    parser.add_argument("--template", help="Sales-admin workbook path for production dry-run mode.")
    parser.add_argument("--production-output", help="Output xlsx path for production write mode.")
    parser.add_argument("--mode", choices=[m.value for m in CollectionMode], default=CollectionMode.ADMIN.value)
    return parser


def login_and_get_stores(client: ServiceClient, args: argparse.Namespace) -> list[Store]:
    """Require a verified authenticated session before any collection begins."""
    while True:
        user_id = args.user_id or input("Service ID: ").strip()
        password = getpass("Password: ")
        try:
            print("[INFO] Logging in...")
            client.login(user_id, password)
            stores = client.get_stores(args.brand_idx)
            if not stores:
                raise AuthenticationError("login verification failed")
            print("[OK] Login succeeded")
            return stores
        except AuthenticationError:
            print("[WARN] Service ID or password is incorrect. Please try again.")
            cookies = getattr(client.session, "cookies", None)
            if cookies is not None:
                cookies.clear()


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        if not args.base_url or not args.brand_idx or not args.brand_name:
            parser.error(
                "--base-url, --brand-idx, and --brand-name are required. "
                "Set COLLECTOR_BASE_URL, COLLECTOR_BRAND_IDX, and COLLECTOR_BRAND_NAME locally."
            )
        if args.production_dry_run and args.production_write:
            parser.error("--production-dry-run and --production-write cannot be used together")
        if args.menu_mapping_preview and not args.menu_monthly_preview:
            args.menu_monthly_preview = True
        if args.menu_excel_dry_run and args.menu_excel_write_copy:
            parser.error("--menu-excel-dry-run and --menu-excel-write-copy cannot be used together")
        if args.menu_excel_write_copy:
            return run_menu_excel_write_copy(args)
        if args.menu_excel_dry_run:
            return run_menu_excel_dry_run(args)
        if args.menu_monthly_preview:
            return run_menu_monthly_preview(args)
        if args.production_dry_run or args.production_write:
            return run_production_mode(args, write=args.production_write)
        if not args.start_date or not args.end_date or not args.output:
            parser.error("--start-date, --end-date, and --output are required outside --production-dry-run")
        start_date = parse_ymd(args.start_date)
        end_date = parse_ymd(args.end_date)
        business_dates = date_range(start_date, end_date)

        client = ServiceClient(args.base_url)
        stores = login_and_get_stores(client, args)
        print(f"[OK] Retrieved {len(stores)} stores")

        selected_stores = resolve_stores(
            stores,
            all_stores=args.all_stores or (not args.store_idx and not args.store_name),
            store_ids=args.store_idx,
            store_names=args.store_name,
        )
        print(f"[INFO] Collecting {len(selected_stores)} stores for {len(business_dates)} day(s)")

        results = collect_sales(
            client,
            business_dates=business_dates,
            brand_idx=args.brand_idx,
            brand_name=args.brand_name,
            stores=selected_stores,
            mode=CollectionMode(args.mode),
        )
        output_path = export_sales(results, args.output, mode=CollectionMode(args.mode))
        print(f"[OK] Exported to {output_path}")
        client.logout()
        return 0
    except ServiceError as exc:
        print(f"[ERROR] {exc}")
        return 1
    except Exception as exc:
        print(f"[ERROR] {exc}")
        return 1


def run_production_mode(args: argparse.Namespace, *, write: bool) -> int:
    if not args.template:
        raise ValueError("--template is required for production mode")
    if args.all_stores or len(args.date) > 1 or (args.store_name and not args.store_idx):
        if args.all_stores and (args.store_idx or args.store_name):
            raise ValueError("use either --all-stores or specific store selection")
        if args.store_idx and not args.store_name:
            raise ValueError("--store-idx requires a matching --store-name")
        if not args.all_stores and not args.store_name:
            raise ValueError("select --all-stores or provide --store-name")
        return run_store_batch_production(args, write=write)
    if len(args.store_idx) != 1 or len(args.store_name) != 1:
        raise ValueError("single-store production requires one --store-idx and one --store-name")
    business_date = parse_ymd(args.date[0]) if args.date else default_target_date()
    client = ServiceClient(args.base_url)
    try:
        print(f"[INFO] Production {'write' if write else 'dry-run'} target: {business_date.isoformat()}")
        login_and_get_stores(client, args)
        record = SingleDaySalesCollector(
            client, brand_idx=args.brand_idx, brand_name=args.brand_name
        ).collect(
            store=Store(args.store_idx[0], args.store_name[0]),
            business_date=business_date,
        )
        preview = SalesAdminDryRun(args.template).preview(record)
        print(f"[PREVIEW] status={preview.status.value}")
        print(f"[PREVIEW] store={record.store_name} date={record.business_date.isoformat()}")
        print(f"[PREVIEW] receipt_count={record.receipt_count} gross_sales_amount={record.gross_sales_amount}")
        for change in preview.changes:
            print(f"[PREVIEW] {change.sheet}!{change.cell} {change.metric}: {change.old_value!r} -> {change.new_value}")
        if preview.reason:
            print(f"[PREVIEW] reason={preview.reason}")
        if not write:
            return 0 if preview.status.value in {"READY", "SAME_VALUE"} else 1
        if preview.status.value != "READY":
            return 1
        if input("Proceed? [y/N] ").strip().lower() != "y":
            print("[INFO] Write cancelled")
            return 0
        output = Path(args.production_output) if args.production_output else Path("output") / f"sales_collection_{business_date:%Y%m%d}.xlsx"
        result = SalesAdminSingleDayWriter(args.template).write(preview, output)
        print(f"[WRITE] output={result.output_path}")
        print(f"[WRITE] cells_written={result.changes_written} original_unchanged={result.original_hash_before == result.original_hash_after}")
        return 0
    finally:
        client.logout()


def run_menu_monthly_preview(args: argparse.Namespace) -> int:
    if not args.year or not args.month:
        raise ValueError("--year and --month are required for --menu-monthly-preview")
    if args.month < 1 or args.month > 12:
        raise ValueError("--month must be between 1 and 12")
    if args.all_stores or args.store_idx:
        raise ValueError("--menu-monthly-preview supports one --store-name only in STEP M-01")
    if len(args.store_name) != 1:
        raise ValueError("--menu-monthly-preview requires exactly one --store-name")

    period_start = date(args.year, args.month, 1)
    period_end = date(args.year, args.month, calendar.monthrange(args.year, args.month)[1])
    client = ServiceClient(args.base_url)
    try:
        print(f"[INFO] Menu monthly preview target: {period_start.isoformat()} ~ {period_end.isoformat()}")
        available_stores = login_and_get_stores(client, args)
        stores = resolve_stores(
            available_stores,
            all_stores=False,
            store_ids=(),
            store_names=args.store_name,
        )
        if not stores:
            raise ValueError("STORE_NOT_FOUND")
        if len(stores) > 1:
            raise ValueError("AMBIGUOUS_STORE_NAME")
        store = stores[0]
        result = client.get_menu_monthly_sales(
            start_date=period_start,
            end_date=period_end,
            brand_idx=args.brand_idx,
            brand_name=args.brand_name,
            store_idx=store.magic_store_id,
            store_name=store.store_name,
        )
        parsed_sum = sum(record.sales_amount for record in result.records)
        parsed_quantity_sum = sum(record.sales_quantity or 0 for record in result.records)
        total_match = "NOT_AVAILABLE" if result.source_total_sales is None else ("YES" if parsed_sum == result.source_total_sales else "NO")
        quantity_match = "NOT_AVAILABLE" if result.source_total_quantity is None else ("YES" if parsed_quantity_sum == result.source_total_quantity else "NO")
        print("=" * 120)
        print("MENU MONTHLY PREVIEW")
        print("=" * 120)
        print(f"STORE_ID={result.store_id}")
        print(f"STORE_NAME={result.store_name}")
        print(f"PERIOD_START={result.period_start.isoformat()}")
        print(f"PERIOD_END={result.period_end.isoformat()}")
        print("SOURCE_STATUS=SUCCESS")
        print("-" * 120)
        print("MENU")
        print("-" * 120)
        for idx, record in enumerate(result.records, 1):
            qty = "N/A" if record.sales_quantity is None else f"{record.sales_quantity:,}"
            unit_price = "N/A" if record.unit_price is None else f"{record.unit_price:,}"
            print(f"{idx:02d} | {record.menu_name} | UNIT={unit_price} | QTY={qty} | SALES={record.sales_amount:,}")
        print("-" * 120)
        print(f"MENU_ROW_COUNT={len(result.records)}")
        print(f"PARSED_QUANTITY_SUM={parsed_quantity_sum:,}")
        print(f"SOURCE_QUANTITY_TOTAL={'NOT_AVAILABLE' if result.source_total_quantity is None else f'{result.source_total_quantity:,}'}")
        print(f"QUANTITY_MATCH={quantity_match}")
        print(f"PARSED_SALES_SUM={parsed_sum:,}")
        print(f"SOURCE_TOTAL={'NOT_AVAILABLE' if result.source_total_sales is None else f'{result.source_total_sales:,}'}")
        print(f"TOTAL_MATCH={total_match}")
        print("=" * 120)
        if args.menu_mapping_preview:
            print_menu_mapping_preview(result)
        return 0
    finally:
        client.logout()


def print_menu_mapping_preview(result) -> None:
    preview = build_menu_mapping_preview(result)
    mapped_sales = preview.sales_by_status(MenuMappingStatus.MAPPED)
    ambiguous_sales = preview.sales_by_status(MenuMappingStatus.AMBIGUOUS)
    unmapped_sales = preview.sales_by_status(MenuMappingStatus.UNMAPPED)
    option_sales = preview.sales_by_status(MenuMappingStatus.NOT_APPLICABLE)
    coverage = (mapped_sales / preview.total_classified_sales * 100) if preview.total_classified_sales else 0

    print("=" * 120)
    print("MENU MAPPING PREVIEW")
    print("=" * 120)
    print(f"STORE={result.store_name}")
    print(f"PERIOD={result.period_start:%Y-%m}")
    print(f"RAW_ROW_COUNT={preview.raw_row_count}")
    print(f"MENU_ROW_COUNT={preview.menu_row_count}")
    print(f"OPTION_ROW_COUNT={preview.option_row_count}")
    print(f"MAPPED_COUNT={preview.count_by_status(MenuMappingStatus.MAPPED)}")
    print(f"AMBIGUOUS_COUNT={preview.count_by_status(MenuMappingStatus.AMBIGUOUS)}")
    print(f"UNMAPPED_COUNT={preview.count_by_status(MenuMappingStatus.UNMAPPED)}")
    print(f"OPTION_COUNT={preview.count_by_status(MenuMappingStatus.NOT_APPLICABLE)}")
    print(f"MAPPED_SALES={mapped_sales:,}")
    print(f"AMBIGUOUS_SALES={ambiguous_sales:,}")
    print(f"UNMAPPED_SALES={unmapped_sales:,}")
    print(f"OPTION_SALES={option_sales:,}")
    print(f"TOTAL_CLASSIFIED_SALES={preview.total_classified_sales:,}")
    print(f"SOURCE_TOTAL_SALES={'NOT_AVAILABLE' if result.source_total_sales is None else f'{result.source_total_sales:,}'}")
    print(f"MAPPING_SALES_COVERAGE={coverage:.2f}%")
    print("-" * 120)
    print("CANONICAL AGGREGATION")
    print("-" * 120)
    for idx, aggregate in enumerate(preview.aggregates, 1):
        print(f"{idx:02d} | {aggregate.canonical_code} | ALIASES={', '.join(aggregate.aliases)} | QTY={aggregate.quantity:,} | SALES={aggregate.sales_amount:,}")
    print("-" * 120)
    print("RAW MENU MAPPING INVENTORY")
    print("-" * 120)
    for idx, item in enumerate(preview.mappings, 1):
        canonical = item.canonical_code or "NONE"
        qty = "N/A" if item.record.sales_quantity is None else f"{item.record.sales_quantity:,}"
        print(f"{idx:02d}.")
        print(f"RAW={item.record.menu_name}")
        print(f"NORMALIZED={item.normalized_name}")
        print("GROUP=UNKNOWN")
        print(f"ROW_TYPE={item.row_type.value}")
        print(f"QTY={qty}")
        print(f"SALES={item.record.sales_amount:,}")
        print(f"CANONICAL={canonical}")
        print(f"STATUS={item.status.value}")
        print(f"REASON={item.reason}")
    print("=" * 120)


def run_menu_excel_dry_run(args: argparse.Namespace) -> int:
    if not args.template:
        raise ValueError("--template is required for --menu-excel-dry-run")
    if not args.year or not args.month:
        raise ValueError("--year and --month are required for --menu-excel-dry-run")
    if args.month < 1 or args.month > 12:
        raise ValueError("--month must be between 1 and 12")
    if args.all_stores or args.store_idx:
        raise ValueError("--menu-excel-dry-run supports one --store-name only in STEP M-04")
    if len(args.store_name) != 1:
        raise ValueError("--menu-excel-dry-run requires exactly one --store-name")

    workbook = load_workbook(args.template, data_only=False)
    profile = infer_menu_template_profile(workbook)
    worksheet = workbook[profile.sheet_name]
    period_start = date(args.year, args.month, 1)
    period_end = date(args.year, args.month, calendar.monthrange(args.year, args.month)[1])
    client = ServiceClient(args.base_url)
    try:
        print(f"[INFO] Menu Excel dry-run target: {period_start.isoformat()} ~ {period_end.isoformat()}")
        available_stores = login_and_get_stores(client, args)
        stores = resolve_stores(available_stores, all_stores=False, store_ids=(), store_names=args.store_name)
        if not stores:
            raise ValueError("STORE_NOT_FOUND")
        if len(stores) > 1:
            raise ValueError("AMBIGUOUS_STORE_NAME")
        store = stores[0]
        source = client.get_menu_monthly_sales(
            start_date=period_start,
            end_date=period_end,
            brand_idx=args.brand_idx,
            brand_name=args.brand_name,
            store_idx=store.magic_store_id,
            store_name=store.store_name,
        )
        mapping_preview = build_menu_mapping_preview(source)
        plan = build_menu_excel_dry_run_plan(mapping_preview, worksheet, profile, store.store_name)
        breakdown = summarize_other_residual_by_reason(plan)

        quantity_sales_row = plan.store_row

        quantity_row = (
            plan.store_row + 1
            if isinstance(plan.store_row, int)
            and plan.store_row > 0
            else 0
        )

        quantity_ratio_row = (
            plan.store_row + 2
            if isinstance(plan.store_row, int)
            and plan.store_row > 0
            else 0
        )

        quantity_source_total = (
            plan.source.source.source_total_quantity
        )

        quantity_direct_total = (
            plan.direct_target_quantity
        )

        quantity_other_total = (
            plan.source_other_residual_quantity
        )

        quantity_option_total = (
            plan.option_quantity
        )

        quantity_by_canonical = {
            aggregate.canonical_code: aggregate.quantity
            for aggregate in plan.source.aggregates
        }

        quantity_business_reconciliation = (
            quantity_direct_total
            + quantity_other_total
            == plan.business_menu_count
        )

        quantity_source_reconciliation = (
            quantity_source_total is not None
            and quantity_direct_total
            + quantity_other_total
            + quantity_option_total
            == quantity_source_total
        )

        print("=" * 120)
        print("MENU EXCEL DRY-RUN")
        print("=" * 120)
        print(f"PROFILE={profile.name}")
        print(f"SHEET={profile.sheet_name}")
        print(f"STORE={store.store_name}")
        print(f"ROW={plan.store_row or 'STORE_NOT_FOUND'}")
        print(f"PERIOD={period_start:%Y-%m}")
        print(f"SOURCE_TOTAL={_money(plan.source_total_sales)}")
        print(f"SOURCE_TOTAL_QUANTITY={'NOT_AVAILABLE' if source.source_total_quantity is None else f'{source.source_total_quantity:,}'}")
        print(f"CURRENT_AC={plan.ac_plan.current_value!r}")
        print(f"PROPOSED_AC={_money(plan.ac_plan.proposed_value)}")
        print(f"AC_STATUS={plan.ac_plan.status.value}")
        print(f"DIRECT_TARGET_TOTAL={_money(plan.direct_target_sales)}")
        print(f"DIRECT_TARGET_QUANTITY={plan.direct_target_quantity:,}")
        print(f"SOURCE_OTHER_RESIDUAL={_money(plan.source_other_residual)}")
        print(f"OTHER_RESIDUAL_QUANTITY={plan.source_other_residual_quantity:,}")
        print(f"OPTION_QUANTITY={plan.option_quantity:,}")
        print(f"RAW_PRODUCT_COUNT={plan.raw_product_count:,}")
        print(f"BUSINESS_MENU_COUNT={plan.business_menu_count:,}")
        print(f"PROPOSED_OTHER_RESIDUAL={_money(plan.calculated_other_residual)}")
        print(f"RESIDUAL_MATCH={'YES' if plan.residual_match else 'NO'}")
        print(f"AB_FORMULA={plan.ab_validation.current_formula}")
        print(f"AB_FORMULA_VALID={'YES' if plan.ab_validation.formula_valid else 'NO'}")
        print(f"AB_WRITE_PLAN_COUNT={plan.ab_write_plan_count}")
        print(f"AD_WRITE_PLAN_COUNT={plan.ad_write_plan_count}")
        print("-" * 120)
        print("QUANTITY ROW PLAN")
        print("-" * 120)

        print(f"SALES_ROW={quantity_sales_row}")
        print(f"QUANTITY_ROW={quantity_row}")
        print(
            f"RATIO_ROW_AFTER_INSERT="
            f"{quantity_ratio_row}"
        )
        print(
            f"QUANTITY_LABEL_CELL=F{quantity_row}"
        )
        print("QUANTITY_LABEL=??")

        for cell in plan.cells:
            print(
                f"COLUMN={cell.target_column} "
                f"CANONICAL={cell.canonical_code} "
                f"CELL={cell.target_column}{quantity_row} "
                f"QUANTITY="f"{quantity_by_canonical.get(cell.canonical_code, 0):,}"
            )

        print(
            f"OTHER_QUANTITY_CELL=AB{quantity_row}"
        )
        print(
            f"OTHER_QUANTITY_FORMULA="
            f"=AC{quantity_row}"
            f"-SUM(G{quantity_row}:AA{quantity_row})"
        )
        print(
            f"TOTAL_QUANTITY_CELL=AC{quantity_row}"
        )
        print(
            "TOTAL_QUANTITY="
            + (
                "NOT_AVAILABLE"
                if quantity_source_total is None
                else f"{quantity_source_total:,}"
            )
        )
        print(
            f"DIRECT_QUANTITY="
            f"{quantity_direct_total:,}"
        )
        print(
            f"OTHER_QUANTITY="
            f"{quantity_other_total:,}"
        )
        print(
            f"OPTION_QUANTITY_FOR_RECONCILIATION="
            f"{quantity_option_total:,}"
        )

        print(
            "BUSINESS_QUANTITY_RECONCILIATION="
            + (
                "YES"
                if quantity_business_reconciliation
                else "NO"
            )
        )

        print(
            "SOURCE_QUANTITY_RECONCILIATION="
            + (
                "YES"
                if quantity_source_reconciliation
                else "NO"
            )
        )

        print("QUANTITY_ROW_INSERT_PLANNED=YES")
        print("QUANTITY_ROW_WRITE_EXECUTED=NO")
        print("EXCEL_WRITE_EXECUTED=NO")

        print("-" * 120)
        print("CELL PLAN")
        print("-" * 120)
        for cell in plan.cells:
            print(
                f"COLUMN={cell.target_column} HEADER={cell.excel_group}/{cell.excel_header} "
                f"CANONICAL={cell.canonical_code} CELL={cell.target_cell} "
                f"QTY={_canonical_quantity(mapping_preview, cell.canonical_code):,} "
                f"CURRENT={cell.current_value!r} PROPOSED={_money(cell.proposed_value)} STATUS={cell.status.value}"
            )
        print("-" * 120)
        print("OTHER RESIDUAL BREAKDOWN")
        print("-" * 120)
        for key, value in breakdown.items():
            print(f"{key}={_money(value)}")

        inventory = other_new_menu_inventory(plan)

        print("-" * 120)
        print("OTHER_NEW_MENU RAW INVENTORY")
        print("-" * 120)

        inventory_sales = 0
        inventory_quantity = 0

        for index, row in enumerate(inventory, start=1):
            inventory_sales += int(row["sales_amount"])
            inventory_quantity += int(row["quantity"])

            avg = row["average_realized_sales"]

            avg_text = (
                "NOT_AVAILABLE"
                if avg is None
                else f"{avg:,.2f}"
            )

            print(
                f"#{index:02d} "
                f"RAW={row['raw_name']} | "
                f"NORMALIZED={row['normalized_name']} | "
                f"UNIT_PRICE={row['unit_price']!r} | "
                f"QTY={row['quantity']:,} | "
                f"SALES={row['sales_amount']:,} | "
                f"AVG_REALIZED={avg_text} | "
                f"STATUS={row['mapping_status']} | "
                f"CANONICAL={row['canonical_code']} | "
                f"REASON={row['mapping_reason']}"
            )

        print("-" * 120)
        print(f"OTHER_NEW_MENU_ROW_COUNT={len(inventory)}")
        print(f"OTHER_NEW_MENU_QUANTITY={inventory_quantity:,}")
        print(f"OTHER_NEW_MENU_SALES={inventory_sales:,}")
        print(
            "OTHER_NEW_MENU_SALES_MATCH="
            + (
                "YES"
                if inventory_sales == breakdown["OTHER_NEW_MENU"]
                else "NO"
            )
        )

        print("=" * 120)
        return 0 if _menu_dry_run_is_pass(plan) else 1
    finally:
        client.logout()


def run_menu_excel_write_copy(args: argparse.Namespace) -> int:
    if not args.output:
        raise ValueError("--output is required for --menu-excel-write-copy")
    plan_context = _build_menu_excel_plan_from_args(args)
    plan = plan_context["plan"]
    profile = plan_context["profile"]
    print_menu_excel_plan(plan, profile)
    if not _menu_dry_run_is_pass(plan):
        raise ValueError("MENU_EXCEL_DRY_RUN_GATE_FAILED")
    result = MenuMonthlyCopyWriter(args.template).write_copy(
        plan,
        args.output,
        profile,
        year=args.year,
        month=args.month,
    )
    print("=" * 120)
    print("MENU EXCEL COPY WRITE")
    print("=" * 120)
    print(f"SOURCE={result.source_path}")
    print(f"OUTPUT={result.output_path}")
    print(f"STORE={result.store}")
    print(f"PERIOD={result.period}")
    print(f"WRITTEN_CELLS={result.written_cells}")
    print(f"SAME_VALUE_CELLS={result.same_value_cells}")
    print(f"ANALYSIS_ROWS={result.analysis_rows}")
    print(f"SALES_RECONCILIATION={'PASS' if result.sales_reconciliation else 'FAIL'}")
    print(f"QUANTITY_RECONCILIATION={'PASS' if result.quantity_reconciliation else 'FAIL'}")
    print(f"FORMULA_VALIDATION={'PASS' if result.formula_protection else 'FAIL'}")
    print(f"ORIGINAL_UNCHANGED={'YES' if result.original_unchanged else 'NO'}")
    print("=" * 120)
    return 0


def _build_menu_excel_plan_from_args(args: argparse.Namespace) -> dict:
    if not args.template:
        raise ValueError("--template is required for menu Excel mode")
    if not args.year or not args.month:
        raise ValueError("--year and --month are required for menu Excel mode")
    if args.month < 1 or args.month > 12:
        raise ValueError("--month must be between 1 and 12")
    if args.all_stores or args.store_idx:
        raise ValueError("menu Excel mode supports one --store-name only")
    if len(args.store_name) != 1:
        raise ValueError("menu Excel mode requires exactly one --store-name")

    workbook = load_workbook(args.template, data_only=False)
    profile = infer_menu_template_profile(workbook)
    worksheet = workbook[profile.sheet_name]
    period_start = date(args.year, args.month, 1)
    period_end = date(args.year, args.month, calendar.monthrange(args.year, args.month)[1])
    client = ServiceClient(args.base_url)
    try:
        print(f"[INFO] Menu Excel target: {period_start.isoformat()} ~ {period_end.isoformat()}")
        available_stores = login_and_get_stores(client, args)
        stores = resolve_stores(available_stores, all_stores=False, store_ids=(), store_names=args.store_name)
        if not stores:
            raise ValueError("STORE_NOT_FOUND")
        if len(stores) > 1:
            raise ValueError("AMBIGUOUS_STORE_NAME")
        store = stores[0]
        source = client.get_menu_monthly_sales(
            start_date=period_start,
            end_date=period_end,
            brand_idx=args.brand_idx,
            brand_name=args.brand_name,
            store_idx=store.magic_store_id,
            store_name=store.store_name,
        )
        mapping_preview = build_menu_mapping_preview(source)
        plan = build_menu_excel_dry_run_plan(mapping_preview, worksheet, profile, store.store_name)
        return {"plan": plan, "profile": profile}
    finally:
        client.logout()


def print_menu_excel_plan(plan, profile) -> None:
    breakdown = summarize_other_residual_by_reason(plan)
    print("=" * 120)
    print("MENU EXCEL DRY-RUN")
    print("=" * 120)
    print(f"PROFILE={profile.name}")
    print(f"SHEET={profile.sheet_name}")
    print(f"STORE={plan.store_name}")
    print(f"SHARED_NORMALIZED_STORE={plan.normalized_store_name}")
    print(f"EXCEL_STORE={plan.excel_store_name}")
    print(f"STORE_MATCH_SOURCE={plan.store_match_source}")
    print(f"ROW={plan.store_row or 'STORE_NOT_FOUND'}")
    print(f"PERIOD={plan.source.source.period_start:%Y-%m}")
    print(f"SOURCE_TOTAL={_money(plan.source_total_sales)}")
    print(f"SOURCE_TOTAL_QUANTITY={'NOT_AVAILABLE' if plan.source.source.source_total_quantity is None else f'{plan.source.source.source_total_quantity:,}'}")
    print(f"CURRENT_AC={plan.ac_plan.current_value!r}")
    print(f"PROPOSED_AC={_money(plan.ac_plan.proposed_value)}")
    print(f"AC_STATUS={plan.ac_plan.status.value}")
    print(f"DIRECT_TARGET_TOTAL={_money(plan.direct_target_sales)}")
    print(f"DIRECT_TARGET_QUANTITY={plan.direct_target_quantity:,}")
    print(f"SOURCE_OTHER_RESIDUAL={_money(plan.source_other_residual)}")
    print(f"OTHER_RESIDUAL_QUANTITY={plan.source_other_residual_quantity:,}")
    print(f"OPTION_QUANTITY={plan.option_quantity:,}")
    print(f"RAW_PRODUCT_COUNT={plan.raw_product_count:,}")
    print(f"BUSINESS_MENU_COUNT={plan.business_menu_count:,}")
    print(f"PROPOSED_OTHER_RESIDUAL={_money(plan.calculated_other_residual)}")
    print(f"RESIDUAL_MATCH={'YES' if plan.residual_match else 'NO'}")
    print(f"AB_FORMULA={plan.ab_validation.current_formula}")
    print(f"AB_FORMULA_VALID={'YES' if plan.ab_validation.formula_valid else 'NO'}")
    print(f"AB_WRITE_PLAN_COUNT={plan.ab_write_plan_count}")
    print(f"AD_WRITE_PLAN_COUNT={plan.ad_write_plan_count}")
    print("-" * 120)
    print("CELL PLAN")
    print("-" * 120)
    for cell in plan.cells:
        print(
            f"COLUMN={cell.target_column} HEADER={cell.excel_group}/{cell.excel_header} "
            f"CANONICAL={cell.canonical_code} CELL={cell.target_cell} "
            f"QTY={_canonical_quantity(plan.source, cell.canonical_code):,} "
            f"CURRENT={cell.current_value!r} PROPOSED={_money(cell.proposed_value)} STATUS={cell.status.value}"
        )
    print("-" * 120)
    print("OTHER RESIDUAL BREAKDOWN")
    print("-" * 120)
    for key, value in breakdown.items():
        print(f"{key}={_money(value)}")
    print("=" * 120)


def run_store_batch_production(args: argparse.Namespace, *, write: bool) -> int:
    business_dates = [parse_ymd(value) for value in args.date] or [default_target_date()]
    template = Path(args.template)
    workbook = load_workbook(template, data_only=False)
    resolver = SalesAdminTemplateResolver(workbook)
    client = ServiceClient(args.base_url)
    failures = 0
    skipped_inactive = 0
    skipped_no_erp_data = 0
    ready_previews = []
    failure_details: list[str] = []
    try:
        print(f"[INFO] All-store production {'write' if write else 'dry-run'} dates: {', '.join(value.isoformat() for value in business_dates)}")
        available_stores = login_and_get_stores(client, args)
        stores = resolve_stores(
            available_stores,
            all_stores=args.all_stores,
            store_ids=args.store_idx,
            store_names=args.store_name,
        )
        if not stores:
            raise ValueError("STORE_NOT_FOUND")
        if args.store_name and not args.store_idx and len(stores) > 1:
            raise ValueError("AMBIGUOUS_STORE_NAME")
        print(f"[INFO] Selected {len(stores)} of {len(available_stores)} stores")
        collector = SingleDaySalesCollector(client, brand_idx=args.brand_idx, brand_name=args.brand_name)
        previewer = SalesAdminDryRun(template)
        for business_date in business_dates:
            for store in stores:
                try:
                    category = resolver.store_category(business_date, store.magic_store_id, store.store_name)
                    if resolver.is_inactive_category(category):
                        skipped_inactive += 1
                        # Inactive stores have no ERP lookup, but their template cells
                        # must explicitly show that the store was not operating.
                        record = SalesAdminDailyRecord(
                            business_date=business_date,
                            store_id=store.magic_store_id,
                            store_name=store.store_name,
                            receipt_count="-",
                            gross_sales_amount="-",
                        )
                    else:
                        record = collector.collect(store=store, business_date=business_date)
                    preview = previewer.preview(record)
                    cells = ", ".join(change.cell for change in preview.changes)
                    source = "INACTIVE" if resolver.is_inactive_category(category) else "ERP"
                    print(f"[PREVIEW] status={preview.status.value} source={source} date={business_date.isoformat()} store={store.store_name} receipt={record.receipt_count} gross={record.gross_sales_amount} cells={cells}")
                    if preview.status.value == "READY":
                        ready_previews.append(preview)
                    elif preview.status.value != "SAME_VALUE":
                        failures += 1
                        detail = f"date={business_date.isoformat()} store={store.store_name} status={preview.status.value} reason={preview.reason}"
                        failure_details.append(detail)
                        print(f"[FAIL] {detail}")
                except Exception as exc:
                    if str(exc) == "UNMATCHED":
                        skipped_no_erp_data += 1
                        print(f"[SKIP] status=SKIPPED_NO_ERP_DATA date={business_date.isoformat()} store={store.store_name}")
                        continue
                    failures += 1
                    detail = f"date={business_date.isoformat()} store={store.store_name} status=FAILED reason={exc}"
                    failure_details.append(detail)
                    print(f"[PREVIEW] {detail}")
        print(
            f"[SUMMARY] dates={len(business_dates)} stores={len(stores)} "
            f"skipped_inactive={skipped_inactive} "
            f"skipped_no_erp_data={skipped_no_erp_data} "
            f"ready_updates={len(ready_previews)} failures={failures}"
        )
        if failure_details:
            print("[FAILURE SUMMARY]")
            for detail in failure_details:
                print(f"[FAIL] {detail}")
            report_output = Path(args.production_output) if args.production_output else Path("output") / "sales_collection_failure.log"
            report_path = report_output.with_suffix(".failure.log")
            report_path.parent.mkdir(parents=True, exist_ok=True)
            report_path.write_text("\n".join(failure_details) + "\n", encoding="utf-8")
            print(f"[FAILURE REPORT] {report_path}")
        if failures:
            return 1
        if not write:
            return 0
        if not ready_previews:
            print("[INFO] No new values require writing")
            return 0
        if input(f"Proceed with {len(ready_previews)} updates in one copied workbook? [y/N] ").strip().lower() != "y":
            print("[INFO] Write cancelled")
            return 0
        date_label = "_".join(value.strftime("%Y%m%d") for value in business_dates)
        output = Path(args.production_output) if args.production_output else Path("output") / f"sales_collection_{date_label}.xlsx"
        result = SalesAdminSingleDayWriter(template).write_many(ready_previews, output)
        print(f"[WRITE] output={result.output_path}")
        print(f"[WRITE] cells_written={result.changes_written} original_unchanged={result.original_hash_before == result.original_hash_after}")
        return 0
    finally:
        client.logout()


def resolve_stores(
    stores: Sequence[Store],
    *,
    all_stores: bool,
    store_ids: Sequence[str],
    store_names: Sequence[str],
) -> list[Store]:
    if all_stores:
        return list(stores)
    selected: list[Store] = []
    id_set = {value.strip() for value in store_ids if value.strip()}
    name_set = {normalize_store_name(value) for value in store_names if value.strip()}
    for store in stores:
        if id_set and store.magic_store_id in id_set:
            selected.append(store)
        elif name_set and normalize_store_name(store.store_name) in name_set:
            selected.append(store)
    return selected


def infer_menu_template_profile(workbook) -> MenuTemplateProfile:
    if DAEJEON_JULY_PROFILE.sheet_name in workbook.sheetnames:
        return DAEJEON_JULY_PROFILE
    if DAEGU_JULY_PROFILE.sheet_name in workbook.sheetnames:
        return DAEGU_JULY_PROFILE
    raise ValueError("MENU_TEMPLATE_PROFILE_NOT_FOUND")


def _menu_dry_run_is_pass(plan) -> bool:
    if not plan.residual_match:
        return False
    if not plan.ab_validation.formula_valid:
        return False
    if plan.ac_plan.status not in {CellPlanStatus.READY, CellPlanStatus.SAME_VALUE}:
        return False
    allowed = {CellPlanStatus.READY, CellPlanStatus.SAME_VALUE}
    return all(cell.status in allowed for cell in plan.cells)


def _money(value: int | None) -> str:
    return "NOT_AVAILABLE" if value is None else f"{value:,}"


def _canonical_quantity(mapping_preview, canonical_code: str) -> int:
    for aggregate in mapping_preview.aggregates:
        if aggregate.canonical_code == canonical_code:
            return aggregate.quantity
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


