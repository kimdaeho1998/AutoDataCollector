from __future__ import annotations

import hashlib
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..mapping.template_resolver import CellType, SalesAdminTemplateResolver, TargetCells
from ..models import DailySalesRecord


@dataclass(frozen=True)
class Change:
    sheet: str
    cell: str
    old: object
    new: int
    metric: str


class SalesAdminTemplateWriter:
    def __init__(self, template: str | Path) -> None:
        self.template = Path(template)

    def dry_run(self, record: DailySalesRecord, store_id: str, store_name: str) -> tuple[TargetCells, list[Change]]:
        wb = load_workbook(self.template, data_only=False)
        targets = SalesAdminTemplateResolver(wb).resolve_target_cells(record.business_date, store_id, store_name)
        ws = wb[targets.sheet_name]
        gross_sales = record.gross_sales_amount if record.gross_sales_amount is not None else record.sales_amount
        changes = [Change(ws.title, targets.receipt_cell, ws[targets.receipt_cell].value, record.receipt_count or 0, "receipt_count"), Change(ws.title, targets.sales_cell, ws[targets.sales_cell].value, gross_sales or 0, "gross_sales_amount")]
        for change in changes:
            if SalesAdminTemplateResolver.classify_cell(ws[change.cell]) != CellType.INPUT:
                raise ValueError("FORMULA_TARGET_OR_CONFLICT")
        return targets, changes

    def write_single(self, record: DailySalesRecord, store_id: str, store_name: str, output: str | Path) -> Path:
        before = self._hash()
        targets, changes = self.dry_run(record, store_id, store_name)
        output = Path(output); output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.template, output)
        wb = load_workbook(output, data_only=False); ws = wb[targets.sheet_name]
        formula_count = sum(1 for row in ws.iter_rows() for cell in row if cell.data_type == "f")
        merged = set(map(str, ws.merged_cells.ranges))
        for change in changes: ws[change.cell] = change.new
        wb.save(output)
        check = load_workbook(output, data_only=False); verified = check[targets.sheet_name]
        expected_sales = record.gross_sales_amount if record.gross_sales_amount is not None else record.sales_amount
        if verified[targets.receipt_cell].value != record.receipt_count or verified[targets.sales_cell].value != expected_sales or formula_count != sum(1 for row in verified.iter_rows() for cell in row if cell.data_type == "f") or merged != set(map(str, verified.merged_cells.ranges)) or before != self._hash():
            raise ValueError("WORKBOOK_INTEGRITY_FAIL")
        return output

    def _hash(self) -> str:
        return hashlib.sha256(self.template.read_bytes()).hexdigest()
