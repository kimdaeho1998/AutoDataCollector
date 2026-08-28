from __future__ import annotations

import hashlib
import shutil
import os
import subprocess
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Sequence

from openpyxl import load_workbook

from ..mapping.template_resolver import CellType, SalesAdminTemplateResolver
from ..production import DryRunStatus, SingleDayDryRunResult


@dataclass(frozen=True)
class SingleDayWriteResult:
    output_path: Path
    changes_written: int
    original_hash_before: str
    original_hash_after: str


class SalesAdminSingleDayWriter:
    """Copy a workbook and safely apply the two-cell production change set."""

    def __init__(self, source_path: str | Path) -> None:
        self.source_path = Path(source_path)

    def write(self, preview: SingleDayDryRunResult, output_path: str | Path) -> SingleDayWriteResult:
        return self.write_many((preview,), output_path)

    def write_many(
        self, previews: Sequence[SingleDayDryRunResult], output_path: str | Path
    ) -> SingleDayWriteResult:
        if not previews:
            raise ValueError("EMPTY_CHANGE_SET")
        if any(preview.status != DryRunStatus.READY for preview in previews):
            statuses = ",".join(preview.status.value for preview in previews)
            raise ValueError(f"WRITE_REQUIRES_READY:{statuses}")
        if any(len(preview.changes) != 2 for preview in previews):
            raise ValueError("INVALID_CHANGE_SET")
        changes = tuple(change for preview in previews for change in preview.changes)
        targets = {(change.sheet, change.cell) for change in changes}
        if len(targets) != len(changes):
            raise ValueError("DUPLICATE_TARGET_CELL")
        output_path = Path(output_path)
        if output_path.resolve() == self.source_path.resolve():
            raise ValueError("OUTPUT_MUST_DIFFER_FROM_SOURCE")
        if output_path.exists():
            output_path.unlink()
        output_path.parent.mkdir(parents=True, exist_ok=True)

        original_hash_before = self._hash(self.source_path)
        source = load_workbook(self.source_path, data_only=False)
        source_snapshot = self._snapshot(source)
        try:
            shutil.copy2(self.source_path, output_path)
            output = load_workbook(output_path, data_only=False)
            for preview in previews:
                self._validate_fresh_preconditions(output, preview)
            self._apply_changes_with_excel(output_path, changes)

            verified = load_workbook(output_path, data_only=False)
            self._verify_output(verified, changes, source_snapshot)
            original_hash_after = self._hash(self.source_path)
            if original_hash_before != original_hash_after:
                raise ValueError("ORIGINAL_MODIFIED")
            return SingleDayWriteResult(output_path, len(changes), original_hash_before, original_hash_after)
        except Exception:
            if output_path.exists():
                output_path.unlink()
            raise

    def _apply_changes_with_excel(self, output_path: Path, changes: Sequence) -> None:
        """Let Excel save its own workbook structures; openpyxl cannot preserve this template's array caches."""
        assignments = "\n".join(
            "\n".join(
                (
                    f"$cell = $workbook.Worksheets.Item('{self._ps_quote(change.sheet)}').Range('{change.cell}')",
                    "$numberFormat = $cell.NumberFormat",
                    f"$cell.Value2 = {self._ps_value(change.new_value)}",
                    "$cell.NumberFormat = $numberFormat",
                )
            )
            for change in changes
        )
        script = f"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$workbook = $null
try {{
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.AutomationSecurity = 3
    $workbook = $excel.Workbooks.Open('{self._ps_quote(str(output_path.resolve()))}', 0, $false)
    {assignments}
    $workbook.Save()
}} finally {{
    if ($null -ne $workbook) {{ $workbook.Close($false) }}
    $excel.Quit()
}}
"""
        powershell = Path(os.environ.get("SystemRoot", r"C:\Windows")) / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe"
        # Full-store writes exceed the Windows command-line limit when all
        # assignments are passed through -EncodedCommand.
        with NamedTemporaryFile("w", suffix=".ps1", encoding="utf-8-sig", delete=False) as handle:
            script_path = Path(handle.name)
            handle.write(script)
        try:
            result = subprocess.run(
                [str(powershell), "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass", "-File", str(script_path)],
                capture_output=True,
                text=True,
                check=False,
            )
        finally:
            script_path.unlink(missing_ok=True)
        if result.returncode != 0:
            detail = (result.stderr or result.stdout).strip()[:500]
            raise ValueError(f"EXCEL_AUTOMATION_FAILED:{detail}")

    @staticmethod
    def _ps_quote(value: str) -> str:
        return value.replace("'", "''")

    @classmethod
    def _ps_value(cls, value: int | str) -> str:
        return str(value) if isinstance(value, int) else f"'{cls._ps_quote(value)}'"

    def _validate_fresh_preconditions(self, workbook, preview: SingleDayDryRunResult) -> None:
        record = preview.record
        targets = SalesAdminTemplateResolver(workbook).resolve_target_cells(record.business_date, record.store_id, record.store_name)
        if {targets.receipt_cell, targets.sales_cell} != {change.cell for change in preview.changes}:
            raise ValueError("STALE_PREVIEW")
        for change in preview.changes:
            cell = workbook[change.sheet][change.cell]
            if SalesAdminTemplateResolver.classify_cell(cell) == CellType.FORMULA or cell.value != change.old_value:
                raise ValueError("STALE_PREVIEW")

    def _verify_output(self, output, changes: Sequence, source_snapshot: dict) -> None:
        for change in changes:
            if output[change.sheet][change.cell].value != change.new_value:
                raise ValueError("VALUE_MISMATCH")
        output_snapshot = self._snapshot(output)
        if source_snapshot["formula_map"] != output_snapshot["formula_map"]:
            raise ValueError("FORMULA_INTEGRITY_FAIL")
        if source_snapshot["merged_ranges"] != output_snapshot["merged_ranges"]:
            raise ValueError("MERGED_RANGE_CHANGED")
        if source_snapshot["sheet_structure"] != output_snapshot["sheet_structure"]:
            raise ValueError("SHEET_STRUCTURE_CHANGED")
        if source_snapshot["hidden_state"] != output_snapshot["hidden_state"]:
            raise ValueError("HIDDEN_STATE_CHANGED")
        if source_snapshot["freeze_panes"] != output_snapshot["freeze_panes"]:
            raise ValueError("FREEZE_PANES_CHANGED")
        expected = {(change.sheet, change.cell): change.new_value for change in changes}
        changed = {key: value for key, value in output_snapshot["values"].items() if source_snapshot["values"].get(key) != value}
        if changed != expected:
            raise ValueError("UNEXPECTED_CELL_CHANGE")

    @staticmethod
    def _hash(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    @staticmethod
    def _snapshot(workbook) -> dict:
        values, formulas, number_formats = {}, {}, {}
        merged_ranges, hidden_state, freeze_panes = {}, {}, {}
        for sheet in workbook.worksheets:
            merged_ranges[sheet.title] = tuple(sorted(map(str, sheet.merged_cells.ranges)))
            hidden_state[sheet.title] = (sheet.sheet_state, tuple(sorted((key, dim.hidden) for key, dim in sheet.row_dimensions.items() if dim.hidden)), tuple(sorted((key, dim.hidden) for key, dim in sheet.column_dimensions.items() if dim.hidden)))
            freeze_panes[sheet.title] = str(sheet.freeze_panes) if sheet.freeze_panes else None
            for row in sheet.iter_rows():
                for cell in row:
                    key = (sheet.title, cell.coordinate)
                    number_formats[key] = cell.number_format
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        signature = SalesAdminSingleDayWriter._formula_signature(cell.value)
                        values[key] = signature
                        formulas[key] = signature
                    else:
                        values[key] = cell.value
        return {"values": values, "formula_map": formulas, "number_formats": number_formats, "merged_ranges": merged_ranges, "sheet_structure": tuple(workbook.sheetnames), "hidden_state": hidden_state, "freeze_panes": freeze_panes}

    @staticmethod
    def _formula_signature(value) -> tuple[str, str, str | None]:
        """Array formulas are recreated on load, so compare their content rather than object identity."""
        return (
            type(value).__name__,
            str(getattr(value, "text", value)),
            str(getattr(value, "ref", "")) or None,
        )
