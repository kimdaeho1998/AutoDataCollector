from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from enum import Enum

from openpyxl.utils import get_column_letter

from .store_normalizer import StoreMatchStatus, StoreResolution, normalize_store_name


class CellType(str, Enum):
    INPUT = "INPUT"
    FORMULA = "FORMULA"
    STATIC = "STATIC"
    UNKNOWN = "UNKNOWN"


@dataclass(frozen=True)
class TargetCells:
    sheet_name: str
    store: StoreResolution
    receipt_cell: str
    sales_cell: str


class SalesAdminTemplateResolver:
    category_column = 1
    store_column = 3
    inactive_categories = frozenset({"중단", "폐점"})

    def __init__(self, workbook) -> None:
        self.workbook = workbook

    def resolve_month_sheet(self, business_date: date):
        name = f"{business_date.month}\uc6d4"
        if name not in self.workbook.sheetnames:
            raise ValueError("SHEET_NOT_FOUND")
        return self.workbook[name]

    def resolve_store_row(self, sheet, magic_store_id: str, magic_store_name: str) -> StoreResolution:
        normalized = normalize_store_name(magic_store_name)
        candidates = [(row, sheet.cell(row, self.store_column).value) for row in range(1, sheet.max_row + 1)]
        matches = [(row, value) for row, value in candidates if isinstance(value, str) and normalize_store_name(value) == normalized]
        if len(matches) != 1:
            return StoreResolution(magic_store_id, magic_store_name, normalized, None, sheet.title, None, StoreMatchStatus.AMBIGUOUS if matches else StoreMatchStatus.UNMATCHED, None)
        row, value = matches[0]
        return StoreResolution(magic_store_id, magic_store_name, normalized, value, sheet.title, row, StoreMatchStatus.MATCHED, "normalized_exact")

    def resolve_date_header(self, sheet, business_date: date) -> tuple[int, int]:
        target = f"{business_date.month}\uc6d4{business_date.day}\uc77c"
        for row in range(1, min(5, sheet.max_row) + 1):
            for col in range(1, sheet.max_column + 1):
                if str(sheet.cell(row, col).value or "").replace(" ", "") == target:
                    for merged in sheet.merged_cells.ranges:
                        if merged.min_row == row and merged.min_col == col:
                            return merged.min_col, merged.max_col
                    return col, col
        raise ValueError("DATE_NOT_FOUND")

    def resolve_metric_columns(self, sheet, columns: tuple[int, int]) -> tuple[int, int]:
        receipt = sales = None
        for col in range(columns[0], columns[1] + 1):
            label = str(sheet.cell(2, col).value or "").replace(" ", "")
            if label == "\uac74\uc218": receipt = col
            if label == "\ucd1d\ub9e4\ucd9c": sales = col
        if receipt is None or sales is None:
            raise ValueError("METRIC_NOT_FOUND")
        return receipt, sales

    def resolve_target_cells(self, business_date: date, magic_store_id: str, magic_store_name: str) -> TargetCells:
        sheet = self.resolve_month_sheet(business_date)
        store = self.resolve_store_row(sheet, magic_store_id, magic_store_name)
        if store.status != StoreMatchStatus.MATCHED or store.row is None:
            raise ValueError(store.status.value)
        columns = self.resolve_metric_columns(sheet, self.resolve_date_header(sheet, business_date))
        return TargetCells(sheet.title, store, f"{get_column_letter(columns[0])}{store.row}", f"{get_column_letter(columns[1])}{store.row}")

    def store_category(self, business_date: date, magic_store_id: str, magic_store_name: str) -> str | None:
        sheet = self.resolve_month_sheet(business_date)
        store = self.resolve_store_row(sheet, magic_store_id, magic_store_name)
        if store.status != StoreMatchStatus.MATCHED or store.row is None:
            raise ValueError(store.status.value)
        value = sheet.cell(store.row, self.category_column).value
        return str(value).strip() if value is not None else None

    @classmethod
    def is_inactive_category(cls, category: str | None) -> bool:
        return category in cls.inactive_categories

    @staticmethod
    def classify_cell(cell) -> CellType:
        if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
            return CellType.FORMULA
        return CellType.INPUT if cell.value is None else CellType.STATIC
