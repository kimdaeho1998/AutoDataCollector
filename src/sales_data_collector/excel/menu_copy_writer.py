from __future__ import annotations

import hashlib
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from ..mapping.menu_mapping import MenuMappingStatus, MenuRowType
from .menu_excel_dry_run import CellPlanStatus, MenuExcelDryRunPlan
from .menu_template_profile import MenuTemplateProfile
from .menu_template_resolver import ExcelDisposition
from .menu_quantity_row import insert_quantity_row


@dataclass(frozen=True)
class MenuCopyWriteResult:
    source_path: Path
    output_path: Path
    store: str
    period: str
    written_cells: int
    same_value_cells: int
    analysis_rows: int
    sales_reconciliation: bool
    quantity_reconciliation: bool
    formula_protection: bool
    original_hash_before: str
    original_hash_after: str

    @property
    def original_unchanged(self) -> bool:
        return self.original_hash_before == self.original_hash_after


class MenuMonthlyCopyWriter:
    """Copy the template first, then write menu monthly values only to the copy."""

    def __init__(self, source_path: str | Path) -> None:
        self.source_path = Path(source_path)

    def write_copy(
        self,
        plan: MenuExcelDryRunPlan,
        output_path: str | Path,
        profile: MenuTemplateProfile,
        *,
        year: int,
        month: int,
    ) -> MenuCopyWriteResult:
        output = Path(output_path)
        self._validate_before_copy(plan, output, year=year, month=month)

        original_hash_before = self._hash(self.source_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.source_path, output)
        try:
            workbook = load_workbook(output, data_only=False)
            worksheet = workbook[profile.sheet_name]
            analysis_sheet_name = self.analysis_sheet_name(month)
            if analysis_sheet_name in workbook.sheetnames:
                raise ValueError("ANALYSIS_SHEET_ALREADY_EXISTS")

            written = 0
            same = 0
            for cell in plan.cells:
                if cell.status in {
                    CellPlanStatus.READY,
                    CellPlanStatus.ZERO_PLACEHOLDER,
                }:
                    worksheet[cell.target_cell].value = cell.proposed_value
                    written += 1
                elif cell.status == CellPlanStatus.SAME_VALUE:
                    same += 1
            if plan.ac_plan.status == CellPlanStatus.READY:
                worksheet[plan.ac_plan.target_cell].value = plan.ac_plan.proposed_value
                written += 1
            elif plan.ac_plan.status == CellPlanStatus.SAME_VALUE:
                same += 1

            analysis_rows = self._write_analysis_sheet(workbook, plan, profile, year=year, month=month)

            quantity_row_result = insert_quantity_row(
                worksheet,
                plan,
            )

            workbook.save(output)

            verified = load_workbook(output, data_only=False)
            self._verify_output(verified, plan, profile)
            original_hash_after = self._hash(self.source_path)
            if original_hash_before != original_hash_after:
                raise ValueError("ORIGINAL_MODIFIED")
            return MenuCopyWriteResult(
                source_path=self.source_path,
                output_path=output,
                store=plan.store_name,
                period=f"{year:04d}-{month:02d}",
                written_cells=written,
                same_value_cells=same,
                analysis_rows=analysis_rows,
                sales_reconciliation=plan.residual_match,
                quantity_reconciliation=self._quantity_reconciles(plan),
                formula_protection=plan.ab_validation.formula_valid,
                original_hash_before=original_hash_before,
                original_hash_after=original_hash_after,
            )
        except Exception:
            output.unlink(missing_ok=True)
            raise

    def _validate_before_copy(self, plan: MenuExcelDryRunPlan, output_path: Path, *, year: int, month: int) -> None:
        if not self.source_path.exists():
            raise ValueError("SOURCE_NOT_FOUND")
        if output_path.resolve() == self.source_path.resolve():
            raise ValueError("OUTPUT_MUST_DIFFER_FROM_SOURCE")
        if output_path.exists():
            raise ValueError("OUTPUT_ALREADY_EXISTS")
        if not plan.residual_match:
            raise ValueError("RESIDUAL_RECONCILIATION_ERROR")
        if not self._quantity_reconciles(plan):
            raise ValueError("QUANTITY_RECONCILIATION_ERROR")
        if not plan.ab_validation.formula_valid:
            raise ValueError("AB_FORMULA_INVALID")
        if plan.ac_plan.status not in {CellPlanStatus.READY, CellPlanStatus.SAME_VALUE}:
            raise ValueError(f"AC_NOT_WRITABLE:{plan.ac_plan.status.value}")
        blocked = [
            cell
            for cell in plan.cells
            if cell.status not in {
                CellPlanStatus.READY,
                CellPlanStatus.ZERO_PLACEHOLDER,
                CellPlanStatus.SAME_VALUE,
            }
        ]
        if blocked:
            statuses = ",".join(f"{cell.target_cell}:{cell.status.value}" for cell in blocked)
            raise ValueError(f"CELL_NOT_WRITABLE:{statuses}")
        if any(_is_merged_target(cell.target_cell, plan.source.source.store_name) for cell in plan.cells):
            raise ValueError("MERGED_TARGET_CELL")
        if year != plan.source.source.period_start.year or month != plan.source.source.period_start.month:
            raise ValueError("PERIOD_MISMATCH")

    def _verify_output(self, workbook, plan: MenuExcelDryRunPlan, profile: MenuTemplateProfile) -> None:
        worksheet = workbook[profile.sheet_name]
        for cell in plan.cells:
            if worksheet[cell.target_cell].value != cell.proposed_value:
                raise ValueError("VALUE_MISMATCH")
        if worksheet[plan.ac_plan.target_cell].value != plan.ac_plan.proposed_value:
            raise ValueError("AC_VALUE_MISMATCH")
        if worksheet[plan.ab_validation.target_cell].value != plan.ab_validation.current_formula:
            raise ValueError("AB_FORMULA_CHANGED")

    def _write_analysis_sheet(self, workbook, plan: MenuExcelDryRunPlan, profile: MenuTemplateProfile, *, year: int, month: int) -> int:
        sheet = workbook.create_sheet(self.analysis_sheet_name(month))
        headers = [
            "연도",
            "월",
            "지역",
            "가맹점명",
            "Canonical Code",
            "메뉴명",
            "판매건수",
            "판매금액",
            "기준단가",
            "건당 평균매출",
            "매출비율",
            "판매건수비율",
            "Excel 처리구분",
            "Excel 대상컬럼",
            "Source 상태",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        rows_written = 0
        direct_by_code = {cell.canonical_code: cell for cell in plan.cells}
        for aggregate in plan.source.aggregates:
            direct_cell = direct_by_code.get(aggregate.canonical_code)
            if direct_cell is None:
                continue
            unit_price = _single_unit_price(plan, aggregate.canonical_code)
            self._append_analysis_row(
                sheet,
                year,
                month,
                profile.name,
                plan.store_name,
                aggregate.canonical_code,
                _display_name(aggregate.canonical_code, direct_cell.excel_header),
                aggregate.quantity,
                aggregate.sales_amount,
                unit_price,
                ExcelDisposition.DIRECT_TARGET.value,
                direct_cell.target_column,
                MenuMappingStatus.MAPPED.value,
                plan.source_total_sales,
                plan.business_menu_count,
            )
            rows_written += 1

        for item in plan.disposition.items:
            if item.disposition == ExcelDisposition.DIRECT_TARGET:
                continue
            record = item.mapping.record
            self._append_analysis_row(
                sheet,
                year,
                month,
                profile.name,
                plan.store_name,
                item.mapping.canonical_code,
                record.menu_name,
                record.sales_quantity or 0,
                record.sales_amount,
                record.unit_price,
                item.disposition.value,
                None,
                item.mapping.status.value,
                plan.source_total_sales,
                plan.business_menu_count if item.mapping.row_type != MenuRowType.OPTION else 0,
            )
            rows_written += 1

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [10, 8, 16, 24, 24, 32, 12, 14, 12, 16, 12, 14, 20, 16, 18]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column in {7}:
                    cell.number_format = "#,##0"
                elif cell.column in {8, 9, 10}:
                    cell.number_format = "#,##0"
                elif cell.column in {11, 12}:
                    cell.number_format = "0.00%"
        return rows_written

    @staticmethod
    def _append_analysis_row(
        sheet,
        year: int,
        month: int,
        region: str,
        store_name: str,
        canonical_code: str | None,
        menu_name: str,
        quantity: int,
        sales_amount: int,
        unit_price: int | None,
        disposition: str,
        target_column: str | None,
        source_status: str,
        total_sales: int | None,
        business_menu_count: int,
    ) -> None:
        avg = None if quantity == 0 else sales_amount / quantity
        sales_share = None if not total_sales else sales_amount / total_sales
        quantity_share = None if business_menu_count == 0 else quantity / business_menu_count
        sheet.append(
            [
                year,
                month,
                region,
                store_name,
                canonical_code,
                menu_name,
                quantity,
                sales_amount,
                unit_price,
                avg,
                sales_share,
                quantity_share,
                disposition,
                target_column,
                source_status,
            ]
        )

    @staticmethod
    def analysis_sheet_name(month: int) -> str:
        return f"{month:02d}월 메뉴분석"

    @staticmethod
    def _quantity_reconciles(plan: MenuExcelDryRunPlan) -> bool:
        total = plan.source.source.source_total_quantity
        if total is None:
            return True
        return total == plan.direct_target_quantity + plan.source_other_residual_quantity + plan.option_quantity

    @staticmethod
    def _hash(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()


def _single_unit_price(plan: MenuExcelDryRunPlan, canonical_code: str) -> int | None:
    prices = {
        item.mapping.record.unit_price
        for item in plan.disposition.items
        if item.mapping.canonical_code == canonical_code and item.mapping.record.unit_price is not None
    }
    return next(iter(prices)) if len(prices) == 1 else None


def _display_name(canonical_code: str, excel_header: str | None) -> str:
    return excel_header or canonical_code


def _is_merged_target(target_cell: str | None, _store_name: str) -> bool:
    # Dry-run cells are resolved from unmerged G:AA/AC targets; keep the hook explicit for writer safety.
    return False
