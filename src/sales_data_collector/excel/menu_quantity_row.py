from __future__ import annotations

from copy import copy
import re
from dataclasses import dataclass
from typing import Any

from openpyxl.formula import Tokenizer
from openpyxl.utils import (
    get_column_letter,
    range_boundaries,
)


DIRECT_START_COLUMN = 7       # G
DIRECT_END_COLUMN = 27        # AA
OTHER_COLUMN = 28             # AB
TOTAL_COLUMN = 29             # AC
UNTOUCHED_COLUMN = 30         # AD

SALES_LABEL = "\ub9e4\ucd9c"
QUANTITY_LABEL = "\uac74\uc218"
RATIO_LABEL = "\ube44\uc728"


@dataclass(frozen=True)
class QuantityRowInsertResult:
    sales_row: int
    quantity_row: int
    ratio_row: int

    source_total_quantity: int
    direct_quantity: int
    other_quantity: int
    option_quantity: int
    ab_quantity: int

    formula_cells_translated: int
    merged_ranges_adjusted: int

    quantity_label: str
    ab_formula: str

    ad_untouched: bool
    ratio_label_preserved: bool


@dataclass(frozen=True)
class _FormulaSnapshot:
    old_coordinate: str
    new_coordinate: str
    formula: str


@dataclass(frozen=True)
class _MergedRangeSnapshot:
    original: str
    transformed: str


def _label(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _quantity_by_canonical(plan) -> dict[str, int]:
    result: dict[str, int] = {}

    for aggregate in plan.source.aggregates:

        code = aggregate.canonical_code
        quantity = int(aggregate.quantity)

        if code in result:
            raise ValueError(
                f"DUPLICATE_QUANTITY_CANONICAL:{code}"
            )

        result[code] = quantity

    return result


def _capture_moved_formulas(
    worksheet,
    insert_row: int,
) -> tuple[_FormulaSnapshot, ...]:

    snapshots: list[_FormulaSnapshot] = []

    for row in worksheet.iter_rows(
        min_row=insert_row,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):

        for cell in row:

            value = cell.value

            if not (
                isinstance(value, str)
                and value.startswith("=")
            ):
                continue

            old_coordinate = cell.coordinate

            new_coordinate = (
                f"{get_column_letter(cell.column)}"
                f"{cell.row + 1}"
            )

            snapshots.append(
                _FormulaSnapshot(
                    old_coordinate=old_coordinate,
                    new_coordinate=new_coordinate,
                    formula=value,
                )
            )

    return tuple(snapshots)


def _transform_merged_range(
    range_string: str,
    insert_row: int,
) -> str:

    (
        min_col,
        min_row,
        max_col,
        max_row,
    ) = range_boundaries(range_string)

    # Entire range starts at/below insertion:
    # shift the whole range down one row.
    if min_row >= insert_row:

        min_row += 1
        max_row += 1

    # Range crosses the insertion boundary:
    # expand it so the new quantity row remains inside
    # the original logical store block.
    elif min_row < insert_row <= max_row:

        max_row += 1

    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _capture_affected_merges(
    worksheet,
    insert_row: int,
) -> tuple[_MergedRangeSnapshot, ...]:

    result: list[_MergedRangeSnapshot] = []

    ranges = [
        str(item)
        for item in worksheet.merged_cells.ranges
    ]

    for range_string in ranges:

        (
            _min_col,
            min_row,
            _max_col,
            max_row,
        ) = range_boundaries(range_string)

        if max_row < insert_row:
            continue

        transformed = _transform_merged_range(
            range_string,
            insert_row,
        )

        result.append(
            _MergedRangeSnapshot(
                original=range_string,
                transformed=transformed,
            )
        )

    return tuple(result)


def _remove_affected_merges(
    worksheet,
    snapshots: tuple[_MergedRangeSnapshot, ...],
) -> None:

    for snapshot in snapshots:

        worksheet.unmerge_cells(
            snapshot.original
        )


def _restore_affected_merges(
    worksheet,
    snapshots: tuple[_MergedRangeSnapshot, ...],
) -> None:

    for snapshot in snapshots:

        worksheet.merge_cells(
            snapshot.transformed
        )


def _copy_row_style(
    worksheet,
    source_row: int,
    destination_row: int,
) -> None:

    source_dimension = worksheet.row_dimensions[
        source_row
    ]

    destination_dimension = worksheet.row_dimensions[
        destination_row
    ]

    destination_dimension.height = (
        source_dimension.height
    )

    destination_dimension.hidden = (
        source_dimension.hidden
    )

    destination_dimension.outlineLevel = (
        source_dimension.outlineLevel
    )

    for column in range(
        1,
        worksheet.max_column + 1,
    ):

        source = worksheet.cell(
            row=source_row,
            column=column,
        )

        destination = worksheet.cell(
            row=destination_row,
            column=column,
        )

        if source.has_style:
            destination._style = copy(
                source._style
            )

        destination.number_format = (
            source.number_format
        )

        destination.font = copy(
            source.font
        )

        destination.fill = copy(
            source.fill
        )

        destination.border = copy(
            source.border
        )

        destination.alignment = copy(
            source.alignment
        )

        destination.protection = copy(
            source.protection
        )


_A1_REFERENCE_RE = re.compile(
    r"(?<![A-Za-z0-9_])"
    r"(\$?[A-Za-z]{1,3})"
    r"(\$?)"
    r"([1-9][0-9]*)"
    r"(?![A-Za-z0-9_])"
)


def _shift_range_token_for_insert(
    value: str,
    insert_row: int,
) -> str:
    """
    Shift only A1-style row references that point to rows
    at or below the physical insertion point.

    Examples for insert_row=7:

        G6          -> G6
        $AC$6       -> $AC$6

        G7          -> G8
        $AC$7       -> $AC$8

        G8          -> G9
        $AC8        -> $AC9

        G8:AA8      -> G9:AA9
    """

    def replace_reference(match: re.Match[str]) -> str:

        column = match.group(1)
        absolute_row_marker = match.group(2)
        row = int(
            match.group(3)
        )

        if row >= insert_row:
            row += 1

        return (
            f"{column}"
            f"{absolute_row_marker}"
            f"{row}"
        )

    return _A1_REFERENCE_RE.sub(
        replace_reference,
        value,
    )


def _translate_formula_for_insert(
    formula: str,
    insert_row: int,
) -> str:
    """
    Apply Excel row-insertion semantics to references inside
    one formula.

    Only formula RANGE operands are changed.

    This intentionally does NOT translate based on the formula
    cell's own movement from old_coordinate to new_coordinate.
    """

    tokenizer = Tokenizer(
        formula
    )

    for token in tokenizer.items:

        if not (
            token.type == "OPERAND"
            and token.subtype == "RANGE"
        ):
            continue

        token.value = (
            _shift_range_token_for_insert(
                token.value,
                insert_row,
            )
        )

    return tokenizer.render()


def _translate_moved_formulas(
    worksheet,
    snapshots: tuple[_FormulaSnapshot, ...],
    insert_row: int,
) -> int:

    translated_count = 0

    for snapshot in snapshots:

        try:

            translated = (
                _translate_formula_for_insert(
                    snapshot.formula,
                    insert_row,
                )
            )

        except Exception as exc:

            raise ValueError(
                "FORMULA_INSERT_TRANSLATION_FAILED:"
                f"{snapshot.old_coordinate}:"
                f"{snapshot.new_coordinate}:"
                f"{snapshot.formula}"
            ) from exc

        worksheet[
            snapshot.new_coordinate
        ].value = translated

        translated_count += 1

    return translated_count




def _source_total_quantity(plan) -> int:

    quantity = (
        plan.source
        .source
        .source_total_quantity
    )

    if quantity is None:

        raise ValueError(
            "SOURCE_TOTAL_QUANTITY_MISSING"
        )

    return int(quantity)


def _validate_preinsert_contract(
    worksheet,
    plan,
) -> tuple[int, int]:

    if not isinstance(plan.store_row, int):

        raise ValueError(
            "STORE_ROW_NOT_RESOLVED"
        )

    if plan.store_row <= 0:

        raise ValueError(
            f"INVALID_STORE_ROW:{plan.store_row}"
        )

    sales_row = plan.store_row
    ratio_row = sales_row + 1

    sales_label = _label(
        worksheet.cell(
            row=sales_row,
            column=6,
        ).value
    )

    ratio_label = _label(
        worksheet.cell(
            row=ratio_row,
            column=6,
        ).value
    )

    if sales_label != SALES_LABEL:

        raise ValueError(
            "SALES_ROW_LABEL_MISMATCH:"
            f"F{sales_row}:"
            f"{sales_label!r}"
        )

    if ratio_label == QUANTITY_LABEL:

        raise ValueError(
            "QUANTITY_ROW_ALREADY_EXISTS:"
            f"F{ratio_row}"
        )

    if ratio_label != RATIO_LABEL:

        raise ValueError(
            "RATIO_ROW_LABEL_MISMATCH:"
            f"F{ratio_row}:"
            f"{ratio_label!r}"
        )



    return (
        sales_row,
        ratio_row,
    )


def insert_quantity_row(
    worksheet,
    plan,
) -> QuantityRowInsertResult:
    """
    Insert one physical quantity row between the resolved
    sales row and ratio row.

    The operation applies ONLY to the already-created output
    workbook handled by MenuMonthlyCopyWriter.

    It never opens or writes the source workbook itself.
    """

    (
        sales_row,
        original_ratio_row,
    ) = _validate_preinsert_contract(
        worksheet,
        plan,
    )

    quantity_row = sales_row + 1
    ratio_row = sales_row + 2

    source_total_quantity = (
        _source_total_quantity(plan)
    )

    direct_quantity = int(
        plan.direct_target_quantity
    )

    other_quantity = int(
        plan.source_other_residual_quantity
    )

    option_quantity = int(
        plan.option_quantity
    )

    ab_quantity = (
        other_quantity
        + option_quantity
    )

    business_menu_count = int(
        plan.business_menu_count
    )

    if (
        direct_quantity
        + other_quantity
        != business_menu_count
    ):

        raise ValueError(
            "BUSINESS_QUANTITY_RECONCILIATION_FAILED:"
            f"DIRECT={direct_quantity}:"
            f"OTHER={other_quantity}:"
            f"BUSINESS={business_menu_count}"
        )

    if (
        business_menu_count
        + option_quantity
        != source_total_quantity
    ):

        raise ValueError(
            "SOURCE_QUANTITY_RECONCILIATION_FAILED:"
            f"BUSINESS={business_menu_count}:"
            f"OPTION={option_quantity}:"
            f"SOURCE={source_total_quantity}"
        )

    quantity_by_canonical = (
        _quantity_by_canonical(plan)
    )

    direct_sum_from_cells = 0

    for cell in plan.cells:

        quantity = int(
            quantity_by_canonical.get(
                cell.canonical_code,
                0,
            )
        )

        direct_sum_from_cells += quantity

    if direct_sum_from_cells != direct_quantity:

        raise ValueError(
            "DIRECT_QUANTITY_CELL_SUM_FAILED:"
            f"CELLS={direct_sum_from_cells}:"
            f"PLAN={direct_quantity}"
        )

    # Capture structural state BEFORE insertion.
    formula_snapshots = (
        _capture_moved_formulas(
            worksheet,
            quantity_row,
        )
    )

    merge_snapshots = (
        _capture_affected_merges(
            worksheet,
            quantity_row,
        )
    )

    original_ratio_label = _label(
        worksheet.cell(
            row=original_ratio_row,
            column=6,
        ).value
    )

    _remove_affected_merges(
        worksheet,
        merge_snapshots,
    )

    worksheet.insert_rows(
        quantity_row,
        amount=1,
    )

    # New quantity row visually follows the sales row.
    _copy_row_style(
        worksheet,
        source_row=sales_row,
        destination_row=quantity_row,
    )

    translated_count = (
        _translate_moved_formulas(
            worksheet,
            formula_snapshots,
            quantity_row,
        )
    )

    # --------------------------------------------------------------------------------
    # Quantity row values
    # --------------------------------------------------------------------------------

    # Clear copied/non-empty values across A:AD.
    # Style is retained.
    for column in range(
        1,
        UNTOUCHED_COLUMN + 1,
    ):

        worksheet.cell(
            row=quantity_row,
            column=column,
        ).value = None

    # F = 건수
    worksheet.cell(
        row=quantity_row,
        column=6,
    ).value = QUANTITY_LABEL

    # Direct menu columns G:AA.
    #
    # Explicit 0 is used for canonical menu columns
    # with no quantity, matching the numeric nature
    # of the template.
    for column in range(
        DIRECT_START_COLUMN,
        DIRECT_END_COLUMN + 1,
    ):

        worksheet.cell(
            row=quantity_row,
            column=column,
        ).value = 0

    for cell in plan.cells:

        quantity = int(
            quantity_by_canonical.get(
                cell.canonical_code,
                0,
            )
        )

        target_column = cell.target_column

        worksheet[
            f"{target_column}{quantity_row}"
        ].value = quantity

    # AC = source total quantity.
    worksheet.cell(
        row=quantity_row,
        column=TOTAL_COLUMN,
    ).value = source_total_quantity

    # AB = residual/other quantity.
    ab_formula = (
        f"=AC{quantity_row}"
        f"-SUM(G{quantity_row}:AA{quantity_row})"
    )

    worksheet.cell(
        row=quantity_row,
        column=OTHER_COLUMN,
    ).value = ab_formula

    # AD is deliberately left blank/untouched.
    worksheet.cell(
        row=quantity_row,
        column=UNTOUCHED_COLUMN,
    ).value = None

    # Restore merged ranges only after row content/style
    # has been applied.
    _restore_affected_merges(
        worksheet,
        merge_snapshots,
    )

    # --------------------------------------------------------------------------------
    # Post validation
    # --------------------------------------------------------------------------------

    current_sales_label = _label(
        worksheet.cell(
            row=sales_row,
            column=6,
        ).value
    )

    current_quantity_label = _label(
        worksheet.cell(
            row=quantity_row,
            column=6,
        ).value
    )

    current_ratio_label = _label(
        worksheet.cell(
            row=ratio_row,
            column=6,
        ).value
    )

    if current_sales_label != SALES_LABEL:

        raise ValueError(
            "POSTINSERT_SALES_LABEL_CHANGED:"
            f"F{sales_row}:"
            f"{current_sales_label!r}"
        )

    if current_quantity_label != QUANTITY_LABEL:

        raise ValueError(
            "POSTINSERT_QUANTITY_LABEL_INVALID:"
            f"F{quantity_row}:"
            f"{current_quantity_label!r}"
        )

    if current_ratio_label != RATIO_LABEL:

        raise ValueError(
            "POSTINSERT_RATIO_LABEL_CHANGED:"
            f"EXPECTED={original_ratio_label!r}:"
            f"ACTUAL={current_ratio_label!r}"
        )

    if current_ratio_label != "비율":

        raise ValueError(
            "POSTINSERT_RATIO_LABEL_INVALID:"
            f"F{ratio_row}:"
            f"{current_ratio_label!r}"
        )

    actual_total = worksheet.cell(
        row=quantity_row,
        column=TOTAL_COLUMN,
    ).value

    if actual_total != source_total_quantity:

        raise ValueError(
            "POSTINSERT_TOTAL_QUANTITY_MISMATCH:"
            f"ACTUAL={actual_total}:"
            f"EXPECTED={source_total_quantity}"
        )

    actual_ab_formula = worksheet.cell(
        row=quantity_row,
        column=OTHER_COLUMN,
    ).value

    if actual_ab_formula != ab_formula:

        raise ValueError(
            "POSTINSERT_AB_FORMULA_MISMATCH:"
            f"ACTUAL={actual_ab_formula!r}:"
            f"EXPECTED={ab_formula!r}"
        )

    ad_value = worksheet.cell(
        row=quantity_row,
        column=UNTOUCHED_COLUMN,
    ).value

    if ad_value is not None:

        raise ValueError(
            "POSTINSERT_AD_NOT_UNTOUCHED:"
            f"AD{quantity_row}={ad_value!r}"
        )

    actual_direct = 0

    for column in range(
        DIRECT_START_COLUMN,
        DIRECT_END_COLUMN + 1,
    ):

        value = worksheet.cell(
            row=quantity_row,
            column=column,
        ).value

        if isinstance(value, (int, float)):
            actual_direct += int(value)

    if actual_direct != direct_quantity:

        raise ValueError(
            "POSTINSERT_DIRECT_QUANTITY_MISMATCH:"
            f"ACTUAL={actual_direct}:"
            f"EXPECTED={direct_quantity}"
        )

    calculated_other = (
        source_total_quantity
        - actual_direct
    )

    if calculated_other != ab_quantity:

        raise ValueError(
            "POSTINSERT_AB_QUANTITY_MISMATCH:"
            f"ACTUAL={calculated_other}:"
            f"OTHER_MENU={other_quantity}:"
            f"OPTION={option_quantity}:"
            f"EXPECTED={ab_quantity}"
        )

    return QuantityRowInsertResult(
        sales_row=sales_row,
        quantity_row=quantity_row,
        ratio_row=ratio_row,
        source_total_quantity=source_total_quantity,
        direct_quantity=direct_quantity,
        other_quantity=other_quantity,
        option_quantity=option_quantity,
        ab_quantity=ab_quantity,
        formula_cells_translated=translated_count,
        merged_ranges_adjusted=len(
            merge_snapshots
        ),
        quantity_label=current_quantity_label,
        ab_formula=ab_formula,
        ad_untouched=True,
        ratio_label_preserved=True,
    )