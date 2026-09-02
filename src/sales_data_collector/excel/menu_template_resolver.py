from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
import re
from typing import Iterable

from openpyxl.utils import get_column_letter

from ..mapping.menu_mapping import MenuMappingPreview, MenuMappingResult, MenuMappingStatus, MenuRowType
from .menu_template_profile import MenuTemplateProfile


class MenuTargetStatus(str, Enum):
    TARGET_RESOLVED = "TARGET_RESOLVED"
    NO_DIRECT_TARGET = "NO_DIRECT_TARGET"
    AMBIGUOUS_SOURCE = "AMBIGUOUS_SOURCE"
    OPTION_NOT_APPLICABLE = "OPTION_NOT_APPLICABLE"
    TARGET_HEADER_MISSING = "TARGET_HEADER_MISSING"
    TARGET_HEADER_DUPLICATE = "TARGET_HEADER_DUPLICATE"
    FORMULA_PROTECTED = "FORMULA_PROTECTED"
    STORE_NOT_FOUND = "STORE_NOT_FOUND"
    STORE_DUPLICATE = "STORE_DUPLICATE"
    RECONCILIATION_ERROR = "RECONCILIATION_ERROR"


class ExcelDisposition(str, Enum):
    DIRECT_TARGET = "DIRECT_TARGET"
    OTHER_RESIDUAL = "OTHER_RESIDUAL"
    OPTION_NOT_APPLICABLE = "OPTION_NOT_APPLICABLE"
    OPTION_REVIEW_REQUIRED = "OPTION_REVIEW_REQUIRED"


@dataclass(frozen=True)
class ExcelMenuTarget:
    canonical_code: str
    group_header: str
    menu_header: str


@dataclass(frozen=True)
class ResolvedMenuTarget:
    canonical_code: str
    status: MenuTargetStatus
    column_letter: str | None = None
    group_header: str | None = None
    menu_header: str | None = None
    reason: str = ""


@dataclass(frozen=True)
class StoreMenuRow:
    row_index: int
    store_name: str
    status: MenuTargetStatus
    reason: str = ""


@dataclass(frozen=True)
class MenuExcelTargetItem:
    mapping: MenuMappingResult
    status: MenuTargetStatus
    disposition: ExcelDisposition
    column_letter: str | None
    reason: str


@dataclass(frozen=True)
class MenuExcelTargetPreview:
    source: MenuMappingPreview
    items: tuple[MenuExcelTargetItem, ...]

    @property
    def direct_target_sales(self) -> int:
        return sum(item.mapping.record.sales_amount for item in self.items if item.status == MenuTargetStatus.TARGET_RESOLVED)

    @property
    def no_direct_target_sales(self) -> int:
        return sum(item.mapping.record.sales_amount for item in self.items if item.status != MenuTargetStatus.TARGET_RESOLVED)

    @property
    def other_residual_sales(self) -> int:
        return sum(item.mapping.record.sales_amount for item in self.items if item.disposition == ExcelDisposition.OTHER_RESIDUAL)

    @property
    def option_sales(self) -> int:
        return sum(
            item.mapping.record.sales_amount
            for item in self.items
            if item.disposition in (ExcelDisposition.OPTION_NOT_APPLICABLE, ExcelDisposition.OPTION_REVIEW_REQUIRED)
        )

    def count_by_status(self, status: MenuTargetStatus) -> int:
        return sum(1 for item in self.items if item.status == status)

    def sales_by_status(self, status: MenuTargetStatus) -> int:
        return sum(item.mapping.record.sales_amount for item in self.items if item.status == status)

    def count_by_disposition(self, disposition: ExcelDisposition) -> int:
        return sum(1 for item in self.items if item.disposition == disposition)

    def sales_by_disposition(self, disposition: ExcelDisposition) -> int:
        return sum(item.mapping.record.sales_amount for item in self.items if item.disposition == disposition)


@dataclass(frozen=True)
class ResidualReconciliation:
    source_total_sales: int
    direct_target_sales: int
    other_residual_sales: int
    option_sales: int
    expected_other_residual: int
    status: MenuTargetStatus
    reason: str

    @property
    def is_pass(self) -> bool:
        return self.status == MenuTargetStatus.TARGET_RESOLVED


MENU_TARGETS: dict[str, ExcelMenuTarget] = {
    "KIMBAP_5": ExcelMenuTarget("KIMBAP_5", "김밥", "5줄"),
    "KIMBAP_10": ExcelMenuTarget("KIMBAP_10", "김밥", "10줄"),
    "KIMBAP_1": ExcelMenuTarget("KIMBAP_1", "김밥", "1줄"),
    "KIMBAP_WASABI_CRAB_MAYO": ExcelMenuTarget("KIMBAP_WASABI_CRAB_MAYO", "김밥", "와사비크래마요(4줄)"),
    "KIMBAP_SPICY_JINMI": ExcelMenuTarget("KIMBAP_SPICY_JINMI", "김밥", "매콤진미꼬마김밥(4줄)"),
    "KIMBAP_TOFU_SKIN": ExcelMenuTarget("KIMBAP_TOFU_SKIN", "김밥", "유부 꼬마김밥(4줄)"),
    "KIMBAP_YUBU": ExcelMenuTarget("KIMBAP_YUBU", "김밥", "유부 꼬마김밥(4줄)"),
    "KIMBAP_SPICY_FISH_CAKE": ExcelMenuTarget("KIMBAP_SPICY_FISH_CAKE", "김밥", "불어묵꼬마김밥(4줄)"),
    "KIMBAP_BUL_EOMUK": ExcelMenuTarget("KIMBAP_BUL_EOMUK", "김밥", "불어묵꼬마김밥(4줄)"),
    "FISH_CAKE_SOUP": ExcelMenuTarget("FISH_CAKE_SOUP", "어묵탕", "어묵탕"),
    "TTEOKBOKKI_MILD": ExcelMenuTarget("TTEOKBOKKI_MILD", "국물떡볶이", "떡볶이(순)"),
    "TTEOKBOKKI_SPICY": ExcelMenuTarget("TTEOKBOKKI_SPICY", "국물떡볶이", "떡볶이(매)"),
    "JJOLMYEON_MILD": ExcelMenuTarget("JJOLMYEON_MILD", "쫄면", "쫄면(순)"),
    "JJOLMYEON_SPICY": ExcelMenuTarget("JJOLMYEON_SPICY", "쫄면", "쫄면(매)"),
    "SEONBI_UDON": ExcelMenuTarget("SEONBI_UDON", "우동", "선비우동"),
    "KIMCHI_UDON": ExcelMenuTarget("KIMCHI_UDON", "우동", "선비 김치우동"),
    "UDON": ExcelMenuTarget("UDON", "우동", "우동"),
    "RAMEN": ExcelMenuTarget("RAMEN", "라면", "라면"),
    "SAUCE_SRIRACHA_MAYO": ExcelMenuTarget("SAUCE_SRIRACHA_MAYO", "소스", "스리라차"),
    "SAUCE_CHEONGYANG": ExcelMenuTarget("SAUCE_CHEONGYANG", "소스", "청양고추"),
    "SAUCE_MAKHANI_CURRY": ExcelMenuTarget("SAUCE_MAKHANI_CURRY", "소스", "마크니커리"),
    "SAUCE_MARKNI_CURRY": ExcelMenuTarget("SAUCE_MARKNI_CURRY", "소스", "마크니커리"),
    "SAUCE_CHEESE": ExcelMenuTarget("SAUCE_CHEESE", "소스", "치즈"),
    "SIKHYE": ExcelMenuTarget("SIKHYE", "음료", "식혜"),
}

NO_DIRECT_TARGET_CODES = frozenset({"DRINK"})


def resolve_menu_target(worksheet, profile: MenuTemplateProfile, canonical_code: str) -> ResolvedMenuTarget:
    if canonical_code in NO_DIRECT_TARGET_CODES:
        return ResolvedMenuTarget(canonical_code, MenuTargetStatus.NO_DIRECT_TARGET, reason="NO_DIRECT_EXCEL_COLUMN")
    target = MENU_TARGETS.get(canonical_code)
    if target is None:
        return ResolvedMenuTarget(canonical_code, MenuTargetStatus.NO_DIRECT_TARGET, reason="CANONICAL_NOT_IN_TEMPLATE_CONTRACT")

    matches: list[int] = []
    for column in range(profile.menu_start_column, profile.direct_menu_end_column + 1):
        group = _header_text(_effective_cell_value(worksheet, 2, column))
        menu = _header_text(_effective_cell_value(worksheet, 3, column)) or group
        if group == target.group_header and menu == target.menu_header:
            matches.append(column)

    if not matches:
        return ResolvedMenuTarget(
            canonical_code=canonical_code,
            status=MenuTargetStatus.TARGET_HEADER_MISSING,
            group_header=target.group_header,
            menu_header=target.menu_header,
            reason="EXPECTED_HEADER_PAIR_NOT_FOUND",
        )
    if len(matches) > 1:
        return ResolvedMenuTarget(
            canonical_code=canonical_code,
            status=MenuTargetStatus.TARGET_HEADER_DUPLICATE,
            group_header=target.group_header,
            menu_header=target.menu_header,
            reason="EXPECTED_HEADER_PAIR_DUPLICATED",
        )
    column = matches[0]
    return ResolvedMenuTarget(
        canonical_code=canonical_code,
        status=MenuTargetStatus.TARGET_RESOLVED,
        column_letter=get_column_letter(column),
        group_header=target.group_header,
        menu_header=target.menu_header,
        reason="HEADER_PAIR_MATCHED",
    )


def resolve_store_sales_row(worksheet, profile: MenuTemplateProfile, store_name: str) -> StoreMenuRow:
    matches: list[int] = []
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=profile.store_column).value != store_name:
            continue
        if worksheet.cell(row=row, column=profile.sales_marker_column).value != "매출":
            continue
        if worksheet.cell(row=row + 1, column=profile.sales_marker_column).value != "비율":
            continue
        matches.append(row)
    if not matches:
        return StoreMenuRow(0, store_name, MenuTargetStatus.STORE_NOT_FOUND, "STORE_SALES_ROW_NOT_FOUND")
    if len(matches) > 1:
        return StoreMenuRow(0, store_name, MenuTargetStatus.STORE_DUPLICATE, "STORE_SALES_ROW_DUPLICATED")
    return StoreMenuRow(matches[0], store_name, MenuTargetStatus.TARGET_RESOLVED, "STORE_SALES_ROW_MATCHED")


def formula_protected_columns(worksheet, row_index: int, columns: Iterable[int]) -> tuple[str, ...]:
    protected: list[str] = []
    for column in columns:
        value = worksheet.cell(row=row_index, column=column).value
        if isinstance(value, str) and value.startswith("="):
            protected.append(get_column_letter(column))
    return tuple(protected)


def build_menu_excel_target_preview(mapping_preview: MenuMappingPreview, worksheet, profile: MenuTemplateProfile) -> MenuExcelTargetPreview:
    items: list[MenuExcelTargetItem] = []
    cache: dict[str, ResolvedMenuTarget] = {}
    for mapping in mapping_preview.mappings:
        status, disposition, column_letter, reason = _status_for_mapping(mapping, worksheet, profile, cache)
        items.append(
            MenuExcelTargetItem(
                mapping=mapping,
                status=status,
                disposition=disposition,
                column_letter=column_letter,
                reason=reason,
            )
        )
    return MenuExcelTargetPreview(source=mapping_preview, items=tuple(items))


def reconcile_other_residual(preview: MenuExcelTargetPreview, source_total_sales: int | None = None) -> ResidualReconciliation:
    total = source_total_sales if source_total_sales is not None else preview.source.source.source_total_sales
    if total is None:
        return ResidualReconciliation(0, preview.direct_target_sales, preview.other_residual_sales, preview.option_sales, 0, MenuTargetStatus.RECONCILIATION_ERROR, "SOURCE_TOTAL_MISSING")
    expected_other_residual = total - preview.direct_target_sales - preview.option_sales
    if expected_other_residual < 0:
        return ResidualReconciliation(
            source_total_sales=total,
            direct_target_sales=preview.direct_target_sales,
            other_residual_sales=preview.other_residual_sales,
            option_sales=preview.option_sales,
            expected_other_residual=expected_other_residual,
            status=MenuTargetStatus.RECONCILIATION_ERROR,
            reason="DIRECT_TARGET_EXCEEDS_SOURCE_TOTAL",
        )
    if expected_other_residual != preview.other_residual_sales:
        return ResidualReconciliation(
            source_total_sales=total,
            direct_target_sales=preview.direct_target_sales,
            other_residual_sales=preview.other_residual_sales,
            option_sales=preview.option_sales,
            expected_other_residual=expected_other_residual,
            status=MenuTargetStatus.RECONCILIATION_ERROR,
            reason="OTHER_RESIDUAL_MISMATCH",
        )
    return ResidualReconciliation(
        source_total_sales=total,
        direct_target_sales=preview.direct_target_sales,
        other_residual_sales=preview.other_residual_sales,
        option_sales=preview.option_sales,
        expected_other_residual=expected_other_residual,
        status=MenuTargetStatus.TARGET_RESOLVED,
        reason="SOURCE_TOTAL_RECONCILED",
    )


def _status_for_mapping(
    mapping: MenuMappingResult,
    worksheet,
    profile: MenuTemplateProfile,
    cache: dict[str, ResolvedMenuTarget],
) -> tuple[MenuTargetStatus, ExcelDisposition, str | None, str]:
    if mapping.row_type == MenuRowType.OPTION:
        if mapping.record.sales_amount:
            return MenuTargetStatus.OPTION_NOT_APPLICABLE, ExcelDisposition.OPTION_REVIEW_REQUIRED, None, "NON_ZERO_OPTION_ROW"
        return MenuTargetStatus.OPTION_NOT_APPLICABLE, ExcelDisposition.OPTION_NOT_APPLICABLE, None, "OPTION_ROW"
    if mapping.status == MenuMappingStatus.AMBIGUOUS:
        return MenuTargetStatus.AMBIGUOUS_SOURCE, ExcelDisposition.OTHER_RESIDUAL, None, mapping.reason
    if mapping.status == MenuMappingStatus.UNMAPPED:
        return MenuTargetStatus.NO_DIRECT_TARGET, ExcelDisposition.OTHER_RESIDUAL, None, mapping.reason
    if mapping.status != MenuMappingStatus.MAPPED or not mapping.canonical_code:
        return MenuTargetStatus.NO_DIRECT_TARGET, ExcelDisposition.OTHER_RESIDUAL, None, mapping.reason

    resolved = cache.get(mapping.canonical_code)
    if resolved is None:
        resolved = resolve_menu_target(worksheet, profile, mapping.canonical_code)
        cache[mapping.canonical_code] = resolved
    disposition = ExcelDisposition.DIRECT_TARGET if resolved.status == MenuTargetStatus.TARGET_RESOLVED else ExcelDisposition.OTHER_RESIDUAL
    return resolved.status, disposition, resolved.column_letter, resolved.reason


def _effective_cell_value(worksheet, row: int, column: int):
    cell = worksheet.cell(row=row, column=column)
    if cell.value is not None:
        return cell.value
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


def _header_text(value) -> str | None:
    if value is None:
        return None
    normalized = " ".join(str(value).replace("\u00a0", " ").split())
    return re.sub(r"\s+\(", "(", normalized)
